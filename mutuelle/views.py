import os
from django.contrib.auth.decorators import login_required
from django.shortcuts import render
from django.db.models import Sum, Q
from mutuelle.models import Adherent, Section, CotisationMensuelle, CotisationDeces, CotisationExceptionnelle
from datetime import datetime
from django.db.models.functions import TruncMonth
from django.contrib import messages
from django.shortcuts import redirect
from django import forms
from django.utils import timezone
from mutuelle.models import Adherent, CotisationMensuelle, CotisationDeces, CotisationExceptionnelle
from mutuelle.models import ParametreCotisation, CotisationMensuelle
from mutuelle.forms import CotMensuelleMultiMoisForm
from mutuelle.forms import CotMensuelleForm, CotDecesForm, CotExceptionnelleForm
from mutuelle.models import Depense, Section
from mutuelle.forms import DepenseForm
from django.shortcuts import get_object_or_404
import pandas as pd
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied
from mutuelle.forms import AssignRoleForm
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from mutuelle.models import Profile
from .models import Exercice
from .forms import ParametreCotisationForm
from django.views.decorators.cache import never_cache


### --- FONCTIONS ROLES ET SECURITE --- ###
def is_president(user):
    return user.groups.filter(name='president').exists() or user.is_superuser

def is_president_section(user):
    return user.groups.filter(name='presidentdesection').exists() or user.is_superuser

def is_secretaire_general(user):
    return user.groups.filter(name='secretairegeneral').exists() or user.is_superuser

def is_tresorier_general(user):
    return user.groups.filter(name='tresoriergeneral').exists() or user.is_superuser

def is_tresorier_section(user):
    return user.groups.filter(name='tresorierdesection').exists() or user.is_superuser

def user_section_allowed(user, section):
    # Pour chef/trésorier de section : autorisé seulement sur leur section
    if is_president(user) or is_tresorier_general(user):
        return True
    if is_president_section(user) or is_tresorier_section(user):
        # À personnaliser selon ta structure utilisateur/section
        return hasattr(user, 'profile') and user.profile.section == section
    return False

### --- VUES PROTEGEES --- ###


@login_required
@never_cache
@user_passes_test(lambda u: is_president(u) or is_secretaire_general(u) or is_president_section(u) or is_tresorier_general(u) or is_tresorier_section(u))
def dashboard(request):
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')
    section_id = request.GET.get('section')

    # Filtre adapté pour chaque modèle selon le champ date correspondant
    q_mens = q_deces = q_exc = Q()
    if debut and fin:
        try:
            debut_dt = datetime.strptime(debut, "%Y-%m-%d")
            fin_dt = datetime.strptime(fin, "%Y-%m-%d")
            q_mens = Q(date_paiement__gte=debut_dt, date_paiement__lte=fin_dt)
            q_deces = Q(date_paiement__gte=debut_dt, date_paiement__lte=fin_dt)
            q_exc = Q(date_debut__gte=debut_dt, date_debut__lte=fin_dt)
        except Exception:
            pass

    sections = Section.objects.all()
    if section_id:
        sections = sections.filter(id=section_id)

    # Stats globales
    total_h = Adherent.objects.filter(sexe__iexact='H').count()
    total_f = Adherent.objects.filter(sexe__iexact='F').count()
    total_adh = total_h + total_f

    tot_mens = CotisationMensuelle.objects.filter(q_mens).aggregate(s=Sum('montant'))['s'] or 0
    tot_deces = CotisationDeces.objects.filter(q_deces).aggregate(s=Sum('montant'))['s'] or 0
    tot_exc = CotisationExceptionnelle.objects.filter(q_exc).aggregate(s=Sum('montant'))['s'] or 0

    # Tableau par section
    data = []
    section_labels = []
    adherents_par_section = []
    for section in sections:
        adh_h = Adherent.objects.filter(section=section, sexe__iexact='H').count()
        adh_f = Adherent.objects.filter(section=section, sexe__iexact='F').count()
        mens = CotisationMensuelle.objects.filter(section=section).filter(q_mens).aggregate(s=Sum('montant'))['s'] or 0
        deces = CotisationDeces.objects.filter(section=section).filter(q_deces).aggregate(s=Sum('montant'))['s'] or 0
        exc = CotisationExceptionnelle.objects.filter(section=section).filter(q_exc).aggregate(s=Sum('montant'))['s'] or 0
        total_sec = mens + deces + exc
        data.append({
            'section': section.nom,
            'adherents': f"{adh_h + adh_f} ({adh_h}/{adh_f})",
            'cot_mens': mens,
            'cot_deces': deces,
            'cot_exc': exc,
            'total_sec': total_sec,
        })
        section_labels.append(section.nom)
        adherents_par_section.append(adh_h + adh_f)

    # Pour la ligne TOTAL dans le tableau
    data.append({
        'section': 'TOTAL',
        'adherents': f"{total_adh} ({total_h}/{total_f})",
        'cot_mens': tot_mens,
        'cot_deces': tot_deces,
        'cot_exc': tot_exc,
        'total_sec': tot_mens + tot_deces + tot_exc,
    })

    # Graphique évolution cotisations mensuelles par mois (12 mois)
    mens_data = CotisationMensuelle.objects.filter(q_mens)
    evolution = mens_data.annotate(month=TruncMonth('date_paiement')).values('month').annotate(s=Sum('montant')).order_by('month')
    mois_labels = []
    chart_mens = []
    for e in evolution:
        mois = e['month'].strftime('%b %Y')
        mois_labels.append(mois)
        chart_mens.append(e['s'] or 0)

    context = {
        'data': data,
        'sections': Section.objects.all(),
        'chart_mens': chart_mens,              # Evolution par mois
        'mois_labels': mois_labels,            # Libellé des mois pour le graphe
        'section_labels': section_labels,      # Pour camembert par section
        'adherents_par_section': adherents_par_section,
        'total_adh': total_adh,
        'total_h': total_h,
        'total_f': total_f,
        'tot_mens': tot_mens,
        'tot_deces': tot_deces,
        'tot_exc': tot_exc,
        'logo_path': 'logo.png',
        'debut': debut or '',
        'fin': fin or '',
        'section_id': section_id or '',
        'couleur_bordeaux': '#800020',
        'couleur_noir': '#222',
        'couleur_blanc': '#fff',
    }
    return render(request, 'niceadmin/dashboard.html', context)


@login_required
def adherents_list(request):
    adherents = Adherent.objects.all()
    user = request.user
    # Restriction pour président/trésorier de section
    if is_president_section(user) or is_tresorier_section(user):
        if hasattr(user, 'profile') and user.profile.section:
            adherents = adherents.filter(section=user.profile.section)
        else:
            adherents = Adherent.objects.none()
    # Filtres GET
    section_id = request.GET.get('section')
    search = request.GET.get('search')
    etat = request.GET.get('etat')

    adherents = Adherent.objects.all()
    if section_id:
        adherents = adherents.filter(section_id=section_id)
    if search:
        adherents = adherents.filter(Q(nom__icontains=search) | Q(prenoms__icontains=search))
    if etat:
        adherents = adherents.filter(etat=etat)

    sections = Section.objects.all()
    etats = Adherent.ETAT_CHOICES

    context = {
        'adherents': adherents,
        'sections': sections,
        'etats': etats,
        'section_id': section_id or '',
        'search': search or '',
        'etat': etat or '',
    }
    return render(request, 'niceadmin/adherents.html', context)



# Formulaire pour ajouter un adhérent
class AdherentForm(forms.ModelForm):
    class Meta:
        model = Adherent
        fields = ['nom', 'prenoms', 'sexe', 'section', 'telephone1', 'telephone2', 'email', 'date_adhesion', 'etat', 'fonction']
        widgets = {
            'date_adhesion': forms.DateInput(attrs={'type': 'date'}),
        }

@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def adherent_create(request):
    form = AdherentForm(request.POST or None)
    if is_president_section(request.user) or is_tresorier_section(request.user):
    # Empêche de choisir une autre section : restreint le QuerySet du champ
        if hasattr(request.user, 'profile') and request.user.profile.section:
            form.fields['section'].queryset = Section.objects.filter(pk=request.user.profile.section.pk)
            form.fields['section'].initial = request.user.profile.section.pk
            # Optionnel : désactive la sélection (pour affichage seule)
            # form.fields['section'].widget.attrs['readonly'] = True
    if request.method == 'POST':
        form = AdherentForm(request.POST)
        if form.is_valid():
            section = form.cleaned_data.get('section')
            if not user_section_allowed(request.user, section):
                raise PermissionDenied("Vous n'avez pas le droit d'ajouter un adhérent à cette section.")
            form.save()
            messages.success(request, "Adhérent ajouté avec succès !")
            return redirect('adherents_list')
    else:
        form = AdherentForm()
    return render(request, 'niceadmin/adherent_form.html', {'form': form})

# Adhérent update view
login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def adherent_update(request, pk):
    form = AdherentForm(request.POST or None)
    if is_president_section(request.user) or is_tresorier_section(request.user):
    # Empêche de choisir une autre section : restreint le QuerySet du champ
        if hasattr(request.user, 'profile') and request.user.profile.section:
            form.fields['section'].queryset = Section.objects.filter(pk=request.user.profile.section.pk)
            form.fields['section'].initial = request.user.profile.section.pk
            # Optionnel : désactive la sélection (pour affichage seule)
            # form.fields['section'].widget.attrs['readonly'] = True
    adherent = Adherent.objects.get(pk=pk)
    if not user_section_allowed(request.user, adherent.section):
        raise PermissionDenied("Vous n'avez pas le droit de modifier cet adhérent.")
    if adherent.etat == 'Décédé':
        messages.error(request, "Impossible de modifier un adhérent décédé.")
        return redirect('adherents_list')
    if request.method == 'POST':
        form = AdherentForm(request.POST, instance=adherent)
        if form.is_valid():
            form.save()
            messages.success(request, "Adhérent modifié avec succès !")
            return redirect('adherents_list')
    else:
        form = AdherentForm(instance=adherent)
    return render(request, 'niceadmin/adherent_form.html', {'form': form, 'modifier': True})

# Adhérent suspendre view
@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u))
def adherent_suspend(request, pk):
    adherent = Adherent.objects.get(pk=pk)
    if not user_section_allowed(request.user, adherent.section):
        raise PermissionDenied("Accès refusé : section non autorisée.")
    if adherent.etat == 'Actif':
        adherent.etat = 'Suspendu'
        messages.success(request, "Adhérent suspendu avec succès !")
    elif adherent.etat == 'Suspendu':
        adherent.etat = 'Actif'
        messages.success(request, "Adhérent réactivé avec succès !")
    adherent.save()
    return redirect('adherents_list')



# Adhérent décès view
@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u))
def adherent_deces(request, pk):
    adherent = Adherent.objects.get(pk=pk)
    if not user_section_allowed(request.user, adherent.section):
        raise PermissionDenied("Accès refusé : section non autorisée.")
    if adherent.etat != 'Décédé':
        adherent.etat = 'Décédé'
        adherent.save()
        messages.success(request, "Décès de l'adhérent déclaré !")
    else:
        messages.warning(request, "Cet adhérent est déjà déclaré décédé.")
    return redirect('adherents_list')

# Adhérent details view
@login_required
def adherent_detail(request, pk):
    adherent = Adherent.objects.get(pk=pk)
    cot_mensuelles = CotisationMensuelle.objects.filter(adherent=adherent).order_by('-annee', '-mois')
    cot_deces = CotisationDeces.objects.filter(adherent=adherent).order_by('-date_paiement')
    cot_exceptionnelles = CotisationExceptionnelle.objects.filter(adherent=adherent).order_by('-date_debut')
    return render(request, 'niceadmin/adherent_detail.html', {
        'adherent': adherent,
        'cot_mensuelles': cot_mensuelles,
        'cot_deces': cot_deces,
        'cot_exceptionnelles': cot_exceptionnelles,
    })


# ----- Formulaires Cotisations -----

class CotMensuelleForm(forms.ModelForm):
    class Meta:
        model = CotisationMensuelle
        fields = ['adherent', 'mois', 'annee', 'montant', 'date_paiement', 'section', 'exercice']
        widgets = {
            'mois': forms.Select(choices=[
                ('Janvier', 'Janvier'), ('Février', 'Février'), ('Mars', 'Mars'), ('Avril', 'Avril'),
                ('Mai', 'Mai'), ('Juin', 'Juin'), ('Juillet', 'Juillet'), ('Août', 'Août'),
                ('Septembre', 'Septembre'), ('Octobre', 'Octobre'), ('Novembre', 'Novembre'), ('Décembre', 'Décembre')
            ]),
            'annee': forms.NumberInput(attrs={'min': 1900, 'max': 2100, 'maxlength': 4}),
            'date_paiement': forms.DateInput(attrs={'type': 'date'}),
            'montant': forms.NumberInput(attrs={'readonly': 'readonly'}),
        }

class CotDecesForm(forms.ModelForm):
    class Meta:
        model = CotisationDeces
        fields = ['adherent', 'deces', 'montant', 'date_paiement', 'section', 'sexe', 'exercice']
        widgets = {
            'date_paiement': forms.DateInput(attrs={'type': 'date'}),
        }

class CotExceptionnelleForm(forms.ModelForm):
    class Meta:
        model = CotisationExceptionnelle
        fields = ['libelle', 'date_debut', 'adherent', 'montant', 'section', 'projet', 'exercice']
        widgets = {
            'date_debut': forms.DateInput(attrs={'type': 'date'}),
        }


# ----- Vues -----
## Liste des cotisations
@login_required
def cotisations_list(request):
    mensuelles = CotisationMensuelle.objects.all()
    user = request.user

    if is_president_section(user) or is_tresorier_section(user):
        if hasattr(user, 'profile') and user.profile.section:
            mensuelles = mensuelles.filter(section=user.profile.section)
        else:
            mensuelles = CotisationMensuelle.objects.none()
    # Filtres
    section_id = request.GET.get('section')
    adherent_id = request.GET.get('adherent')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')
    type_cotisation = request.GET.get('type')

    mensuelles = CotisationMensuelle.objects.all()
    deces = CotisationDeces.objects.all()
    exceptionnelles = CotisationExceptionnelle.objects.all()

    if section_id:
        mensuelles = mensuelles.filter(section_id=section_id)
        deces = deces.filter(section_id=section_id)
        exceptionnelles = exceptionnelles.filter(section_id=section_id)
    if adherent_id:
        mensuelles = mensuelles.filter(adherent_id=adherent_id)
        deces = deces.filter(adherent_id=adherent_id)
        exceptionnelles = exceptionnelles.filter(adherent_id=adherent_id)
    if debut and fin:
        try:
            from datetime import datetime
            debut_dt = datetime.strptime(debut, "%Y-%m-%d")
            fin_dt = datetime.strptime(fin, "%Y-%m-%d")
            mensuelles = mensuelles.filter(date_paiement__gte=debut_dt, date_paiement__lte=fin_dt)
            deces = deces.filter(date_paiement__gte=debut_dt, date_paiement__lte=fin_dt)
            exceptionnelles = exceptionnelles.filter(date_debut__gte=debut_dt, date_debut__lte=fin_dt)
        except:
            pass

    sections = Section.objects.all()
    adherents = Adherent.objects.all()

    context = {
        'mensuelles': mensuelles.order_by('-annee', '-mois'),
        'deces': deces.order_by('-date_paiement'),
        'exceptionnelles': exceptionnelles.order_by('-date_debut'),
        'sections': sections,
        'adherents': adherents,
        'section_id': section_id or '',
        'adherent_id': adherent_id or '',
        'debut': debut or '',
        'fin': fin or '',
        'type_cotisation': type_cotisation or '',
    }
    return render(request, 'niceadmin/cotisations.html', context)

@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def cotisation_create(request):
    type_cot = request.GET.get('type', 'mensuelle')
    form = None

    if type_cot == 'mensuelle':
        form = CotMensuelleForm(request.POST or None)
        montant_param = ParametreCotisation.objects.filter(type_cotisation='Mensuelle', actif=True).order_by('-date_effet').first()
        if montant_param:
            form.fields['montant'].initial = montant_param.montant

    elif type_cot == 'deces':
        form = CotDecesForm(request.POST or None)

    elif type_cot == 'exceptionnelle':
        form = CotExceptionnelleForm(request.POST or None)

    else:
        messages.error(request, "Type de cotisation inconnu.")
        return redirect('cotisations_list')

    # Sécurité section
    if is_president_section(request.user) or is_tresorier_section(request.user):
        if hasattr(request.user, 'profile') and request.user.profile.section:
            form.fields['section'].queryset = Section.objects.filter(pk=request.user.profile.section.pk)
            form.fields['section'].initial = request.user.profile.section.pk

    if request.method == 'POST':
        if form.is_valid():
            section = form.cleaned_data.get('section')
            exercice = form.cleaned_data.get('exercice')
            if not user_section_allowed(request.user, section):
                raise PermissionDenied("Vous n'avez pas le droit d'ajouter une cotisation pour cette section.")
            if exercice and exercice.cloture:
                messages.error(request, "Impossible d'ajouter : l'exercice sélectionné est clôturé.")
                return redirect('cotisations_list')

            # Vérification anti-doublon
            if type_cot == 'mensuelle':
                adherent = form.cleaned_data['adherent']
                mois = form.cleaned_data['mois']
                annee = form.cleaned_data['annee']
                if CotisationMensuelle.objects.filter(adherent=adherent, mois=mois, annee=annee, exercice=exercice).exists():
                    messages.error(request, "Une cotisation mensuelle existe déjà pour cet adhérent, ce mois, cette année et cet exercice.")
                else:
                    form.save()
                    messages.success(request, "Cotisation mensuelle ajoutée avec succès !")
                    return redirect('cotisations_list')

            elif type_cot == 'deces':
                adherent = form.cleaned_data['adherent']
                deces = form.cleaned_data['deces']
                if CotisationDeces.objects.filter(adherent=adherent, deces=deces, exercice=exercice).exists():
                    messages.error(request, "Cette cotisation décès existe déjà pour cet adhérent, ce décès et cet exercice.")
                else:
                    form.save()
                    messages.success(request, "Cotisation décès ajoutée avec succès !")
                    return redirect('cotisations_list')

            elif type_cot == 'exceptionnelle':
                form.save()
                messages.success(request, "Cotisation exceptionnelle ajoutée avec succès !")
                return redirect('cotisations_list')

    return render(request, 'niceadmin/cotisation_form.html', {'form': form, 'type_cot': type_cot})



@login_required
def cotisation_mensuelle_multimois(request):
    montant_param = ParametreCotisation.objects.filter(type_cotisation='Mensuelle', actif=True).order_by('-date_effet').first()
    montant = montant_param.montant if montant_param else 0
    if request.method == 'POST':
        form = CotMensuelleMultiMoisForm(request.POST)
        form.fields['montant'].initial = montant
        if form.is_valid():
            adherent = form.cleaned_data['adherent']
            annee = form.cleaned_data['annee']
            mois_list = form.cleaned_data['mois']
            date_paiement = form.cleaned_data['date_paiement']
            section = adherent.section
            exercice = form.cleaned_data['exercice']
            if exercice.cloture:
                messages.error(request, "Impossible d'ajouter : l'exercice sélectionné est clôturé.")
                return redirect('cotisations_list')
            doublons = []
            enregistres = []
            for mois in mois_list:
                if CotisationMensuelle.objects.filter(adherent=adherent, mois=mois, annee=annee, exercice=exercice).exists():
                    doublons.append(mois)
                else:
                    CotisationMensuelle.objects.create(
                        adherent=adherent, mois=mois, annee=annee, montant=montant,
                        date_paiement=date_paiement, section=section, exercice=exercice
                    )
                    enregistres.append(mois)
            if enregistres:
                messages.success(request, f"Cotisation(s) enregistrée(s) pour : {', '.join(enregistres)}.")
            if doublons:
                messages.warning(request, f"Déjà existant pour : {', '.join(doublons)}.")
            return redirect('cotisations_list')
    else:
        form = CotMensuelleMultiMoisForm(initial={'montant': montant})
        form.fields['montant'].initial = montant

    return render(request, 'niceadmin/cotisation_mensuelle_form.html', {'form': form, 'montant': montant})

from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.exceptions import PermissionDenied

@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def cotisation_update(request, type, pk):
    form = AdherentForm(request.POST or None)
    if is_president_section(request.user) or is_tresorier_section(request.user):
        # Empêche de choisir une autre section : restreint le QuerySet du champ
        if hasattr(request.user, 'profile') and request.user.profile.section:
            form.fields['section'].queryset = Section.objects.filter(pk=request.user.profile.section.pk)
            form.fields['section'].initial = request.user.profile.section.pk
            # Optionnel : désactive la sélection (pour affichage seule)
            # form.fields['section'].widget.attrs['readonly'] = True

    if type == 'mensuelle':
        Cotisation = CotisationMensuelle
        Form = CotMensuelleForm
    elif type == 'deces':
        Cotisation = CotisationDeces
        Form = CotDecesForm
    elif type == 'exceptionnelle':
        Cotisation = CotisationExceptionnelle
        Form = CotExceptionnelleForm
    else:
        messages.error(request, "Type de cotisation inconnu.")
        return redirect('cotisations_list')

    cot = Cotisation.objects.get(pk=pk)

    # Sécurité section : Seul président/trésorier général ou président/trésorier de section pour SA section
    if not user_section_allowed(request.user, cot.section):
        raise PermissionDenied("Vous n'avez pas le droit de modifier cette cotisation.")

    # Blocage si exercice clôturé
    if cot.exercice.cloture:
        messages.error(request, "Impossible de modifier : l'exercice est clôturé.")
        return redirect('cotisations_list')

    if request.method == 'POST':
        form = Form(request.POST, instance=cot)
        if form.is_valid():
            form.save()
            messages.success(request, "Cotisation modifiée avec succès !")
            return redirect('cotisations_list')
    else:
        form = Form(instance=cot)
    return render(request, 'niceadmin/cotisation_form.html', {'form': form, 'type_cot': type, 'modifier': True})


@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def cotisation_delete(request, type, pk):
    if type == 'mensuelle':
        Cotisation = CotisationMensuelle
    elif type == 'deces':
        Cotisation = CotisationDeces
    elif type == 'exceptionnelle':
        Cotisation = CotisationExceptionnelle
    else:
        messages.error(request, "Type de cotisation inconnu.")
        return redirect('cotisations_list')

    cot = Cotisation.objects.get(pk=pk)

    # Sécurité section : Seul président/trésorier général ou président/trésorier de section pour SA section
    if not user_section_allowed(request.user, cot.section):
        raise PermissionDenied("Vous n'avez pas le droit de supprimer cette cotisation.")

    # Blocage si exercice clôturé
    if cot.exercice.cloture:
        messages.error(request, "Impossible de supprimer : l'exercice est clôturé.")
        return redirect('cotisations_list')

    if request.method == 'POST':
        cot.delete()
        messages.success(request, "Cotisation supprimée avec succès !")
        return redirect('cotisations_list')
    return render(request, 'niceadmin/cotisation_confirm_delete.html', {'cot': cot, 'type_cot': type})



@login_required
def depenses_list(request):
    section_id = request.GET.get('section')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')
    type_depense = request.GET.get('type_depense')

    depenses = Depense.objects.all()
    if section_id:
        depenses = depenses.filter(section_id=section_id)
    if type_depense:
        depenses = depenses.filter(type_depense=type_depense)
    if debut and fin:
        depenses = depenses.filter(date__gte=debut, date__lte=fin)

    sections = Section.objects.all()
    types = Depense.TYPE_CHOICES
    return render(request, 'niceadmin/depenses.html', {
        'depenses': depenses.order_by('-date'),
        'sections': sections,
        'section_id': section_id or '',
        'debut': debut or '',
        'fin': fin or '',
        'types': types,
        'type_depense': type_depense or '',
    })

@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def depense_create(request):
    form = AdherentForm(request.POST or None)
    if is_president_section(request.user) or is_tresorier_section(request.user):
    # Empêche de choisir une autre section : restreint le QuerySet du champ
        if hasattr(request.user, 'profile') and request.user.profile.section:
            form.fields['section'].queryset = Section.objects.filter(pk=request.user.profile.section.pk)
            form.fields['section'].initial = request.user.profile.section.pk
            # Optionnel : désactive la sélection (pour affichage seule)
            # form.fields['section'].widget.attrs['readonly'] = True
    if request.method == 'POST':
        form = DepenseForm(request.POST)
        if form.is_valid():
            section = form.cleaned_data.get('section')
            if not user_section_allowed(request.user, section):
                raise PermissionDenied("Vous n'avez pas le droit d'ajouter une dépense pour cette section.")
            form.save()
            messages.success(request, "Dépense ajoutée avec succès !")
            return redirect('depenses_list')
    else:
        form = DepenseForm()
    return render(request, 'niceadmin/depense_form.html', {'form': form})

@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def depense_update(request, pk):
    form = AdherentForm(request.POST or None)
    if is_president_section(request.user) or is_tresorier_section(request.user):
        # Empêche de choisir une autre section : restreint le QuerySet du champ
        if hasattr(request.user, 'profile') and request.user.profile.section:
            form.fields['section'].queryset = Section.objects.filter(pk=request.user.profile.section.pk)
            form.fields['section'].initial = request.user.profile.section.pk
            # Optionnel : désactive la sélection (pour affichage seule)
            # form.fields['section'].widget.attrs['readonly'] = True

    depense = get_object_or_404(Depense, pk=pk)

    # Contrôle de la section pour chef/ trésorier de section
    if not user_section_allowed(request.user, depense.section):
        raise PermissionDenied("Vous n’avez pas le droit de modifier cette dépense.")

    # Blocage si exercice clôturé
    if depense.exercice.cloture:
        messages.error(request, "Impossible de modifier : l'exercice est clôturé.")
        return redirect('depenses_list')

    if request.method == 'POST':
        form = DepenseForm(request.POST, instance=depense)
        if form.is_valid():
            form.save()
            messages.success(request, "Dépense modifiée avec succès !")
            return redirect('depenses_list')
    else:
        form = DepenseForm(instance=depense)
    return render(request, 'niceadmin/depense_form.html', {'form': form, 'modifier': True})



@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u) or is_president_section(u) or is_tresorier_section(u))
def depense_delete(request, pk):
    depense = get_object_or_404(Depense, pk=pk)
    if not user_section_allowed(request.user, depense.section):
        raise PermissionDenied("Vous n'avez pas le droit de supprimer cette dépense.")
    
    # Blocage si exercice clôturé
    if depense.exercice.cloture:
        messages.error(request, "Impossible de supprimer : l'exercice est clôturé.")
        return redirect('depenses_list')

    if request.method == 'POST':
        depense.delete()
        messages.success(request, "Dépense supprimée avec succès !")
        return redirect('depenses_list')
    return render(request, 'niceadmin/depense_confirm_delete.html', {'depense': depense})



@login_required
def bilan_global(request):
    user = request.user

    if is_president_section(user) or is_tresorier_section(user):
        if hasattr(user, 'profile') and user.profile.section:
            section_id = user.profile.section.id
        else:
            section_id = None
    if section_id:
        cot_mens = cot_mens.filter(section_id=section_id)
        cot_deces = cot_deces.filter(section_id=section_id)
        cot_excep = cot_excep.filter(section_id=section_id)
        depenses = depenses.filter(section_id=section_id)
    # Filtres
    section_id = request.GET.get('section')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')

    sections = Section.objects.all()

    # Cotisations
    cot_mens = CotisationMensuelle.objects.all()
    cot_deces = CotisationDeces.objects.all()
    cot_excep = CotisationExceptionnelle.objects.all()
    depenses = Depense.objects.all()

    # Filtres
    if section_id:
        cot_mens = cot_mens.filter(section_id=section_id)
        cot_deces = cot_deces.filter(section_id=section_id)
        cot_excep = cot_excep.filter(section_id=section_id)
        depenses = depenses.filter(section_id=section_id)
    if debut and fin:
        cot_mens = cot_mens.filter(date_paiement__gte=debut, date_paiement__lte=fin)
        cot_deces = cot_deces.filter(date_paiement__gte=debut, date_paiement__lte=fin)
        cot_excep = cot_excep.filter(date_debut__gte=debut, date_debut__lte=fin)
        depenses = depenses.filter(date__gte=debut, date__lte=fin)

    total_cot_mens = cot_mens.aggregate(total=Sum('montant'))['total'] or 0
    total_cot_deces = cot_deces.aggregate(total=Sum('montant'))['total'] or 0
    total_cot_excep = cot_excep.aggregate(total=Sum('montant'))['total'] or 0
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or 0

    solde = total_cot_mens + total_cot_deces + total_cot_excep - total_depenses

    context = {
        'sections': sections,
        'section_id': section_id or '',
        'debut': debut or '',
        'fin': fin or '',
        'total_cot_mens': total_cot_mens,
        'total_cot_deces': total_cot_deces,
        'total_cot_excep': total_cot_excep,
        'total_depenses': total_depenses,
        'solde': solde,
    }
    return render(request, 'niceadmin/bilan.html', context)



@login_required
@user_passes_test(lambda u: is_president(u) or is_secretaire_general(u) or is_president_section(u) or is_tresorier_general(u) or is_tresorier_section(u))
def bilan_global(request):
    section_id = request.GET.get('section')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')

    sections = Section.objects.all()

    # QuerySets filtrés
    cot_mens = CotisationMensuelle.objects.all()
    cot_deces = CotisationDeces.objects.all()
    cot_excep = CotisationExceptionnelle.objects.all()
    depenses = Depense.objects.all()

    if section_id:
        cot_mens = cot_mens.filter(section_id=section_id)
        cot_deces = cot_deces.filter(section_id=section_id)
        cot_excep = cot_excep.filter(section_id=section_id)
        depenses = depenses.filter(section_id=section_id)
    if debut and fin:
        cot_mens = cot_mens.filter(date_paiement__gte=debut, date_paiement__lte=fin)
        cot_deces = cot_deces.filter(date_paiement__gte=debut, date_paiement__lte=fin)
        cot_excep = cot_excep.filter(date_debut__gte=debut, date_debut__lte=fin)
        depenses = depenses.filter(date__gte=debut, date__lte=fin)

    total_cot_mens = cot_mens.aggregate(total=Sum('montant'))['total'] or 0
    total_cot_deces = cot_deces.aggregate(total=Sum('montant'))['total'] or 0
    total_cot_excep = cot_excep.aggregate(total=Sum('montant'))['total'] or 0
    total_depenses = depenses.aggregate(total=Sum('montant'))['total'] or 0

    solde = total_cot_mens + total_cot_deces + total_cot_excep - total_depenses

    context = {
        'sections': sections,
        'section_id': section_id or '',
        'debut': debut or '',
        'fin': fin or '',
        'total_cot_mens': total_cot_mens,
        'total_cot_deces': total_cot_deces,
        'total_cot_excep': total_cot_excep,
        'total_depenses': total_depenses,
        'solde': solde,
    }
    return render(request, 'niceadmin/bilan.html', context)



@login_required
def bilan_export_excel(request):
    section_id = request.GET.get('section')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')

    # Collecte données
    data = []
    cot_mens = CotisationMensuelle.objects.all()
    cot_deces = CotisationDeces.objects.all()
    cot_excep = CotisationExceptionnelle.objects.all()
    depenses = Depense.objects.all()

    if section_id:
        cot_mens = cot_mens.filter(section_id=section_id)
        cot_deces = cot_deces.filter(section_id=section_id)
        cot_excep = cot_excep.filter(section_id=section_id)
        depenses = depenses.filter(section_id=section_id)
    if debut and fin:
        cot_mens = cot_mens.filter(date_paiement__gte=debut, date_paiement__lte=fin)
        cot_deces = cot_deces.filter(date_paiement__gte=debut, date_paiement__lte=fin)
        cot_excep = cot_excep.filter(date_debut__gte=debut, date_debut__lte=fin)
        depenses = depenses.filter(date__gte=debut, date__lte=fin)

    for c in cot_mens:
        data.append({'Type':'Cotisation Mensuelle', 'Date':c.date_paiement, 'Montant':c.montant, 'Section':c.section.nom})
    for c in cot_deces:
        data.append({'Type':'Cotisation Décès', 'Date':c.date_paiement, 'Montant':c.montant, 'Section':c.section.nom})
    for c in cot_excep:
        data.append({'Type':'Cotisation Exceptionnelle', 'Date':c.date_debut, 'Montant':c.montant, 'Section':c.section.nom})
    for d in depenses:
        data.append({'Type':'Dépense', 'Date':d.date, 'Montant':-d.montant, 'Section':d.section.nom if d.section else ''})
    df = pd.DataFrame(data)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=bilan_mutuelle.xlsx'
    df.to_excel(response, index=False)
    return response





@login_required
def bilan_drilldown(request):
    t = request.GET.get('type')
    section_id = request.GET.get('section')
    debut = request.GET.get('debut')
    fin = request.GET.get('fin')
    objs = []
    label = ''
    if t == "mensuelle":
        qs = CotisationMensuelle.objects.all()
        label = "Cotisations Mensuelles"
    elif t == "deces":
        qs = CotisationDeces.objects.all()
        label = "Cotisations Décès"
    elif t == "exceptionnelle":
        qs = CotisationExceptionnelle.objects.all()
        label = "Cotisations Exceptionnelles"
    elif t == "depense":
        qs = Depense.objects.all()
        label = "Dépenses"
    else:
        qs = []

    if section_id:
        qs = qs.filter(section_id=section_id)
    if debut and fin:
        field = 'date_paiement'
        if t == "exceptionnelle":
            field = 'date_debut'
        elif t == "depense":
            field = 'date'
        qs = qs.filter(**{f"{field}__gte": debut, f"{field}__lte": fin})

    objs = list(qs)
    html = render_to_string("niceadmin/bilan_drilldown_table.html", {'objs':objs, 'type':t, 'label':label})
    return HttpResponse(html)


def is_admin_or_president(user):
    return user.is_superuser or user.groups.filter(name='president').exists()



@user_passes_test(is_admin_or_president)
def assign_role(request):
    # Traitement formulaire assignation
    if request.method == 'POST':
        form = AssignRoleForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data['user']
            group = form.cleaned_data['group']
            section = form.cleaned_data['section']
            # Retire tous les anciens groupes "rôles"
            user.groups.clear()
            user.groups.add(group)
            # Met à jour la section du profil si elle existe
            if hasattr(user, 'profile'):
                user.profile.section = section
                user.profile.save()
            messages.success(request, "Rôle et section mis à jour pour %s." % user.username)
            return redirect('assign_role')
    else:
        form = AssignRoleForm()

    # Liste tous les users avec groupe & section pour affichage
    User = get_user_model()
    users = User.objects.filter(is_superuser=False).order_by('username')
    users_info = []
    for u in users:
        groupe = u.groups.first().name if u.groups.exists() else '-'
        section = u.profile.section.nom if hasattr(u, 'profile') and u.profile.section else '-'
        users_info.append({'id': u.id, 'username': u.username, 'email': u.email, 'groupe': groupe, 'section': section})

    return render(request, 'niceadmin/gestion_roles.html', {
        'form': form,
        'users_info': users_info,
    })





def is_president_ou_tresorier(user):
    return user.groups.filter(name__in=['president', 'tresorier']).exists() or user.is_superuser

@login_required
def cloturer_exercice(request, pk):
    exercice = get_object_or_404(Exercice, pk=pk)
    if exercice.cloture:
        messages.info(request, "Cet exercice est déjà clôturé.")
        return redirect('exercices_list')  # ou ta page liste
    if request.method == "POST":
        exercice.cloture = True
        exercice.save()
        messages.success(request, "Exercice clôturé avec succès !")
        return redirect('exercices_list')
    return render(request, 'niceadmin/cloturer_exercice_confirm.html', {'exercice': exercice})

def is_super_admin(user):
    return user.is_superuser

@login_required
@user_passes_test(is_super_admin)
def decloturer_exercice(request, pk):
    exercice = get_object_or_404(Exercice, pk=pk)
    if not exercice.cloture:
        messages.info(request, "Cet exercice n'est pas clôturé.")
        return redirect('exercices_list')
    if request.method == "POST":
        exercice.cloture = False
        exercice.save()
        messages.success(request, "Exercice réouvert (déclôturé) avec succès !")
        return redirect('exercices_list')
    return render(request, 'niceadmin/decloturer_exercice_confirm.html', {'exercice': exercice})

@login_required
def exercices_list(request):
    exercices = Exercice.objects.all()
    return render(request, 'niceadmin/exercices_list.html', {'exercices': exercices})


from .forms import ExerciceForm

@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name__in=['president', 'tresorier']).exists())
def exercice_create(request):
    if request.method == 'POST':
        form = ExerciceForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Exercice créé avec succès !")
            return redirect('exercices_list')
    else:
        form = ExerciceForm()
    return render(request, 'niceadmin/exercice_create.html', {'form': form})


@login_required
def profile_view(request):
    return render(request, 'niceadmin/profile.html', {'user': request.user})



@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u))
def parametres_cotisation_list(request):
    params = ParametreCotisation.objects.filter(type_cotisation='Mensuelle').order_by('-date_effet')

    if request.method == 'POST':
        form = ParametreCotisationForm(request.POST)
        if form.is_valid():
            param = form.save(commit=False)
            param.type_cotisation = 'Mensuelle'  # Forçage ici
            param.save()
            messages.success(request, "Paramètre enregistré avec succès.")
            return redirect('parametres_cotisation_list')
    else:
        form = ParametreCotisationForm()

    return render(request, 'niceadmin/parametres_cotisation.html', {
        'form': form,
        'params': params,
    })


@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u))
def parametre_toggle_actif(request, pk):
    param = get_object_or_404(ParametreCotisation, pk=pk)
    param.actif = not param.actif
    param.save()
    return redirect('parametres_cotisation_list')


@login_required
@user_passes_test(lambda u: is_president(u) or is_tresorier_general(u))
def parametre_edit(request, pk):
    param = get_object_or_404(ParametreCotisation, pk=pk)

    if request.method == 'POST':
        form = ParametreCotisationForm(request.POST, instance=param)
        if form.is_valid():
            instance = form.save(commit=False)
            instance.type_cotisation = "Mensuelle"  # toujours forcé
            instance.save()
            messages.success(request, "Paramètre modifié.")
            return redirect('parametres_cotisation_list')
    else:
        form = ParametreCotisationForm(instance=param)

    return render(request, 'niceadmin/parametre_edit.html', {
        'form': form,
        'param': param
    })






import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from mutuelle.models import Section  # déjà présent normalement

@login_required
def export_membres_excel(request):
    section_id = request.GET.get("section")
    sexe = request.GET.get("sexe")
    etat = request.GET.get("etat")

    membres = Adherent.objects.all()

    if section_id:
        membres = membres.filter(section_id=section_id)
    if sexe:
        membres = membres.filter(sexe=sexe)
    if etat:
        membres = membres.filter(etat=etat)

    section_nom = "Toutes sections"
    if section_id:
        try:
            section = Section.objects.get(id=section_id)
            section_nom = section.nom
        except Section.DoesNotExist:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Membres"

    # ✅ Logo
    logo_path = os.path.join(settings.BASE_DIR, 'mutuelle/static/logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.height = 60
        img.width = 60
        ws.add_image(img, "A1")

    # ✅ Titre dynamique
    titre = f"Liste des Membres – Mutuelle de Développement de Djangoménou"
    sous_titre = f"Section : {section_nom}"

    ws.merge_cells("A2:F2")
    ws["A2"] = titre
    ws["A2"].font = Font(size=14, bold=True, color="800000")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:F3")
    ws["A3"] = sous_titre
    ws["A3"].font = Font(size=12, italic=True)
    ws["A3"].alignment = Alignment(horizontal="center")

    # ✅ Entête
    entetes = ["Nom", "Prénoms", "Sexe", "Section", "Téléphone", "Email"]
    ws.append(entetes)

    for cell in ws[4]:  # ligne 4
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # ✅ Données
    for m in membres:
        ws.append([
            m.nom,
            m.prenoms,
            m.sexe,
            m.section.nom if m.section else "",
            m.telephone1,
            m.email or ""
        ])

    # ✅ Auto-ajustement largeur colonnes
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # ✅ Réponse HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="liste_membres.xlsx"'
    wb.save(response)
    return response




from xhtml2pdf import pisa # type: ignore
from django.template.loader import render_to_string
from django.http import HttpResponse
from io import BytesIO
from django.conf import settings

@login_required
def export_membres_pdf(request):
    section_id = request.GET.get("section")
    membres = Adherent.objects.all()
    section_nom = "Toutes sections"

    if section_id:
        try:
            section = Section.objects.get(id=section_id)
            membres = membres.filter(section=section)
            section_nom = section.nom
        except Section.DoesNotExist:
            pass  # Garde "Toutes sections"

    html = render_to_string("exports/membres_pdf.html", {
        'membres': membres,
        'section_nom': section_nom,
        'logo_path': f"{settings.BASE_DIR}/mutuelle/static/logo.png"
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename=liste_membres.pdf'
    pisa.CreatePDF(BytesIO(html.encode("utf-8")), dest=response)
    return response


from mutuelle.models import CotisationMensuelle, CotisationDeces, CotisationExceptionnelle

@login_required
def export_cotisations_excel(request):
    section_id = request.GET.get('section')
    type_cotisation = request.GET.get('type')  # mensuelle, deces, exceptionnelle
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')

    cotisations = []
    section_nom = "Toutes sections"

    if type_cotisation == "mensuelle":
        cotisations = CotisationMensuelle.objects.all()
    elif type_cotisation == "deces":
        cotisations = CotisationDeces.objects.all()
    elif type_cotisation == "exceptionnelle":
        cotisations = CotisationExceptionnelle.objects.all()
    else:
        # tout type
        cotisations = list(CotisationMensuelle.objects.all()) + list(CotisationDeces.objects.all()) + list(CotisationExceptionnelle.objects.all())

    # Filtrage par section
    if section_id:
        cotisations = [c for c in cotisations if hasattr(c, 'adherent') and c.adherent.section_id == int(section_id)]
        from mutuelle.models import Section
        try:
            section_nom = Section.objects.get(id=section_id).nom
        except:
            pass

    # Filtrage par période
    if date_debut and date_fin:
        from datetime import datetime
        try:
            d1 = datetime.strptime(date_debut, '%Y-%m-%d').date()
            d2 = datetime.strptime(date_fin, '%Y-%m-%d').date()
            cotisations = [c for c in cotisations if d1 <= c.date_paiement <= d2]
        except:
            pass

    # Export Excel
@login_required
def export_cotisations_excel(request):
    section_id = request.GET.get('section')
    type_cotisation = request.GET.get('type')
    date_debut = request.GET.get('date_debut')
    date_fin = request.GET.get('date_fin')
    adherent_id = request.GET.get('adherent')

    from mutuelle.models import Section
    from datetime import datetime

    cotisations = []
    section_nom = "Toutes sections"
    adherent_nom = "Tous les adhérents"

    if type_cotisation == "mensuelle":
        cotisations = CotisationMensuelle.objects.all()
    elif type_cotisation == "deces":
        cotisations = CotisationDeces.objects.all()
    elif type_cotisation == "exceptionnelle":
        cotisations = CotisationExceptionnelle.objects.all()
    else:
        cotisations = list(CotisationMensuelle.objects.all()) + list(CotisationDeces.objects.all()) + list(CotisationExceptionnelle.objects.all())

    if section_id:
        cotisations = [c for c in cotisations if hasattr(c, 'section') and c.section.id == int(section_id)]
        try:
            section_nom = Section.objects.get(id=section_id).nom
        except:
            pass

    if adherent_id:
        cotisations = [c for c in cotisations if hasattr(c, 'adherent') and c.adherent.id == int(adherent_id)]
        try:
            adh = Adherent.objects.get(id=adherent_id)
            adherent_nom = f"{adh.nom} {adh.prenoms}"
        except:
            pass

    if date_debut and date_fin:
        try:
            d1 = datetime.strptime(date_debut, "%Y-%m-%d").date()
            d2 = datetime.strptime(date_fin, "%Y-%m-%d").date()
            cotisations = [c for c in cotisations if hasattr(c, 'date_paiement') and d1 <= c.date_paiement <= d2 or hasattr(c, 'date_debut') and d1 <= c.date_debut <= d2]
        except:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cotisations"

    # ✅ Titre dynamique
    titre = f"Cotisations – Type: {type_cotisation or 'Tous'} – Section: {section_nom} – Adhérent: {adherent_nom}"
    ws.merge_cells("A1:F1")
    ws["A1"] = titre
    ws["A1"].font = Font(size=13, bold=True, color="800000")
    ws["A1"].alignment = Alignment(horizontal="center")

    # ✅ En-têtes
    headers = ["Nom", "Prénoms", "Type", "Montant", "Date", "Section"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="800000", end_color="800000", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # ✅ Données
    for c in cotisations:
        nom = c.adherent.nom if hasattr(c, 'adherent') else ''
        prenoms = c.adherent.prenoms if hasattr(c, 'adherent') else ''
        montant = int(c.montant)
        section = c.section.nom if hasattr(c, 'section') else ''
        if hasattr(c, 'date_paiement'):
            date_str = c.date_paiement.strftime('%d/%m/%Y')
        elif hasattr(c, 'date_debut'):
            date_str = c.date_debut.strftime('%d/%m/%Y')
        else:
            date_str = ''
        cot_type = type(c).__name__.replace("Cotisation", "")
        ws.append([nom, prenoms, cot_type, montant, date_str, section])

    # ✅ Ajustement automatique
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="cotisations_filtrees.xlsx"'
    wb.save(response)
    return response


from xhtml2pdf import pisa # type: ignore
from django.template.loader import get_template
from django.http import HttpResponse
from django.conf import settings
import os
from django.utils.timezone import now
from datetime import datetime

@login_required
def export_cotisations_pdf(request):
    type_cotisation = request.GET.get('type', '')
    section_id = request.GET.get('section', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    cotisations = CotisationMensuelle.objects.select_related('adherent__section').all()
    section_nom = "Toutes sections"

    if type_cotisation == "deces":
        cotisations = CotisationDeces.objects.select_related('adherent__section').all()
    elif type_cotisation == "exceptionnelle":
        cotisations = CotisationExceptionnelle.objects.select_related('adherent__section').all()

    if section_id:
        cotisations = cotisations.filter(adherent__section_id=section_id)
        section = Section.objects.filter(id=section_id).first()
        if section:
            section_nom = section.nom

    if date_debut:
        cotisations = cotisations.filter(date_paiement__gte=date_debut)
    if date_fin:
        cotisations = cotisations.filter(date_paiement__lte=date_fin)
# Calcule du total
    total = sum(c.montant for c in cotisations)
    template = get_template('exports/cotisations_pdf.html')
    html = template.render({
        'cotisations': cotisations,
        'logo_path': os.path.join(settings.BASE_DIR, 'mutuelle', 'static', 'logo.png'),
        'type': type_cotisation.capitalize() if type_cotisation else "Toutes",
        'section': section_nom,
        'date': now().strftime("%d/%m/%Y"),
        'periode': f"{date_debut} au {date_fin}" if date_debut and date_fin else "",
        'total': total,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="cotisations.pdf"'
    pisa.CreatePDF(src=html, dest=response)
    return response



import pandas as pd
from mutuelle.models import Adherent, Section
from django.contrib import messages
from django.shortcuts import render, redirect

@login_required
def import_adherents(request):
    rapport = ""
    if request.method == "POST" and request.FILES.get("fichier"):
        fichier = request.FILES["fichier"]
        try:
            # Détecte le type de fichier
            if fichier.name.endswith(".csv"):
                df = pd.read_csv(fichier)
            else:
                df = pd.read_excel(fichier)

            count = 0
            erreurs = 0
            for i, row in df.iterrows():
                try:
                    section_obj = Section.objects.get(nom=row["section"])
                    adherent, created = Adherent.objects.get_or_create(
                        nom=row["nom"],
                        prenoms=row["prenoms"],
                        sexe=row["sexe"],
                        telephone1=row["telephone"],
                        email=row["email"],
                        section=section_obj
                    )
                    if created:
                        count += 1
                except Exception as e:
                    erreurs += 1

            rapport = f"{count} adhérents ajoutés avec succès. {erreurs} erreurs détectées."
        except Exception as e:
            rapport = f"Erreur lors de l’importation : {e}"

    return render(request, "niceadmin/import_adherents.html", {"rapport": rapport})




import pandas as pd
from mutuelle.models import Section, Adherent, CotisationMensuelle, CotisationDeces, CotisationExceptionnelle

@login_required
def import_cotisations(request):
    rapport = ""
    if request.method == "POST" and request.FILES.get("fichier"):
        fichier = request.FILES["fichier"]
        try:
            if fichier.name.endswith(".csv"):
                df = pd.read_csv(fichier)
            else:
                df = pd.read_excel(fichier)

            total = 0
            erreurs = 0

            for i, row in df.iterrows():
                try:
                    nom = row["nom"]
                    prenoms = row["prenoms"]
                    type_cot = row["type"].lower()
                    montant = float(row["montant"])
                    date_paiement = pd.to_datetime(row["date_paiement"]).date()
                    section_nom = row["section"]

                    section = Section.objects.get(nom=section_nom)
                    adherent = Adherent.objects.get(nom=nom, prenoms=prenoms, section=section)

                    if type_cot == "mensuelle":
                        CotisationMensuelle.objects.create(adherent=adherent, montant=montant, date_paiement=date_paiement)
                    elif type_cot == "deces":
                        CotisationDeces.objects.create(adherent=adherent, montant=montant, date_paiement=date_paiement)
                    elif type_cot == "exceptionnelle":
                        CotisationExceptionnelle.objects.create(adherent=adherent, montant=montant, date_paiement=date_paiement)
                    else:
                        raise ValueError("Type de cotisation invalide")

                    total += 1
                except Exception as e:
                    erreurs += 1

            rapport = f"{total} cotisations importées avec succès. {erreurs} lignes en erreur."
        except Exception as e:
            rapport = f"Erreur : {e}"

    return render(request, "niceadmin/import_cotisations.html", {"rapport": rapport})



import pandas as pd
from mutuelle.models import Depense, Section

@login_required
def import_depenses(request):
    rapport = ""
    if request.method == "POST" and request.FILES.get("fichier"):
        fichier = request.FILES["fichier"]
        try:
            if fichier.name.endswith(".csv"):
                df = pd.read_csv(fichier)
            else:
                df = pd.read_excel(fichier)

            total = 0
            erreurs = 0

            for _, row in df.iterrows():
                try:
                    libelle = row["libelle"]
                    montant = float(row["montant"])
                    date_depense = pd.to_datetime(row["date_depense"]).date()
                    section_nom = row["section"]

                    section = Section.objects.get(nom=section_nom)
                    Depense.objects.create(libelle=libelle, montant=montant, date=date_depense, section=section)
                    total += 1
                except Exception as e:
                    erreurs += 1

            rapport = f"{total} dépenses ajoutées. {erreurs} lignes ignorées pour erreur."
        except Exception as e:
            rapport = f"Erreur lors de l'import : {e}"

    return render(request, "niceadmin/import_depenses.html", {"rapport": rapport})


@login_required
def profile_view(request):
    return render(request, 'niceadmin/profile.html')




from django.contrib.auth.models import User, Group
from django.contrib import messages
from mutuelle.models import Section
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required

@login_required
def ajout_utilisateur(request):
    sections = Section.objects.all()

    if request.method == "POST":
        username = request.POST.get("username")
        first_name = request.POST.get("first_name")
        email = request.POST.get("email")
        password = request.POST.get("password")
        password_confirm = request.POST.get("password_confirm")
        role = request.POST.get("role")
        is_admin = request.user.username == "aby"
        section_id = request.POST.get("section")

        # Vérification mots de passe
        if password != password_confirm:
            messages.error(request, "Les mots de passe ne correspondent pas.")
            return render(request, "niceadmin/ajout_utilisateur.html", {"sections": sections})

        # Création de l'utilisateur
        if User.objects.filter(username=username).exists():
            messages.error(request, "Ce nom d'utilisateur existe déjà.")
            return render(request, "niceadmin/ajout_utilisateur.html", {"sections": sections})

        user = User.objects.create_user(username=username, email=email, password=password, first_name=first_name)

        # Attribution du groupe
        if role == "admin" and request.user.username != "aby":
            user.delete()
            messages.error(request, "Vous n'avez pas l'autorisation de créer un utilisateur admin.")
            return render(request, "niceadmin/ajout_utilisateur.html", {"sections": sections, "is_admin": False})

        # Attribution du groupe
        if role:
            group, created = Group.objects.get_or_create(name=role)
            user.groups.add(group)

        # Affectation section si applicable
        if section_id:
            from mutuelle.models import Profile  # si tu as un modèle Profile
            section = Section.objects.get(id=section_id)
            profile = Profile.objects.get(user=user)
            profile.section = section
            profile.save()

        messages.success(request, "Utilisateur ajouté avec succès.")
        return redirect("ajout_utilisateur")

    return render(request, "niceadmin/ajout_utilisateur.html", {"sections": sections, "is_admin": request.user.username == "aby"})
